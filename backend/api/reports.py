"""
Report generation utilities: PDF and Excel export.
"""
import io
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def _header_style():
    return {
        'BACKGROUND': colors.HexColor('#1a5276'),
        'TEXTCOLOR': colors.white,
        'FONTNAME': 'Helvetica-Bold',
        'FONTSIZE': 10,
        'ALIGN': 'CENTER',
        'VALIGN': 'MIDDLE',
        'BOTTOMPADDING': 8,
        'TOPPADDING': 8,
    }


def generate_athletes_pdf(athletes):
    """Generate PDF report of all athletes."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, spaceAfter=20)
    elements.append(Paragraph('Athlete Performance & Injury Tracking System', title_style))
    elements.append(Paragraph(f'Athletes Report - Generated: {datetime.now().strftime("%d-%m-%Y %H:%M")}', styles['Normal']))
    elements.append(Spacer(1, 20))

    data = [['ID', 'Name', 'Sport', 'Team', 'Gender', 'Status', 'Phone', 'Email']]
    for a in athletes:
        data.append([
            str(a.id), a.full_name, a.sport, a.team or '-',
            a.gender, a.status, a.phone or '-', a.email or '-'
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a5276')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f3f4')]),
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_performance_pdf(records):
    """Generate PDF report of performance records."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    elements = []
    styles = getSampleStyleSheet()

    elements.append(Paragraph('Performance Report', styles['Heading1']))
    elements.append(Paragraph(f'Generated: {datetime.now().strftime("%d-%m-%Y %H:%M")}', styles['Normal']))
    elements.append(Spacer(1, 20))

    data = [['Athlete', 'Date', 'Speed', 'Strength', 'Endurance', 'Flexibility', 'Agility']]
    for r in records:
        data.append([
            r.athlete.full_name,
            r.record_date.strftime('%d-%m-%Y'),
            str(r.speed_score or '-'),
            str(r.strength_score or '-'),
            str(r.endurance_score or '-'),
            str(r.flexibility_score or '-'),
            str(r.agility_score or '-'),
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a5276')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f3f4')]),
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_injuries_pdf(injuries):
    """Generate PDF report of injury records."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    elements.append(Paragraph('Injury Report', styles['Heading1']))
    elements.append(Spacer(1, 20))

    data = [['Athlete', 'Type', 'Body Part', 'Date', 'Severity', 'Status']]
    for i in injuries:
        data.append([
            i.athlete.full_name, i.injury_type, i.body_part,
            i.injury_date.strftime('%d-%m-%Y'), i.severity, i.recovery_status
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#c0392b')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer


def _excel_header_style():
    return Font(bold=True, color='FFFFFF', size=11)


def _excel_header_fill():
    return PatternFill(start_color='1A5276', end_color='1A5276', fill_type='solid')


def _excel_border():
    thin = Side(style='thin')
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def generate_athletes_excel(athletes):
    """Export athletes data to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Athletes'

    headers = ['ID', 'First Name', 'Last Name', 'Email', 'Phone', 'Sport', 'Team', 'Gender', 'Status', 'DOB']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _excel_header_style()
        cell.fill = _excel_header_fill()
        cell.alignment = Alignment(horizontal='center')
        cell.border = _excel_border()

    for row_idx, a in enumerate(athletes, 2):
        ws.cell(row=row_idx, column=1, value=a.id)
        ws.cell(row=row_idx, column=2, value=a.first_name)
        ws.cell(row=row_idx, column=3, value=a.last_name)
        ws.cell(row=row_idx, column=4, value=a.email)
        ws.cell(row=row_idx, column=5, value=a.phone)
        ws.cell(row=row_idx, column=6, value=a.sport)
        ws.cell(row=row_idx, column=7, value=a.team)
        ws.cell(row=row_idx, column=8, value=a.gender)
        ws.cell(row=row_idx, column=9, value=a.status)
        ws.cell(row=row_idx, column=10, value=str(a.date_of_birth))

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 15

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_performance_excel(records):
    """Export performance data to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Performance'

    headers = ['Athlete', 'Date', 'Speed Score', 'Strength Score', 'Endurance Score',
               'Flexibility Score', 'Agility Score', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _excel_header_style()
        cell.fill = _excel_header_fill()

    for row_idx, r in enumerate(records, 2):
        ws.cell(row=row_idx, column=1, value=r.athlete.full_name)
        ws.cell(row=row_idx, column=2, value=str(r.record_date))
        ws.cell(row=row_idx, column=3, value=float(r.speed_score) if r.speed_score else '')
        ws.cell(row=row_idx, column=4, value=float(r.strength_score) if r.strength_score else '')
        ws.cell(row=row_idx, column=5, value=float(r.endurance_score) if r.endurance_score else '')
        ws.cell(row=row_idx, column=6, value=float(r.flexibility_score) if r.flexibility_score else '')
        ws.cell(row=row_idx, column=7, value=float(r.agility_score) if r.agility_score else '')
        ws.cell(row=row_idx, column=8, value=r.notes)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_attendance_excel(records):
    """Export attendance data to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Attendance'

    headers = ['Athlete', 'Date', 'Status', 'Session Type', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _excel_header_style()
        cell.fill = _excel_header_fill()

    for row_idx, r in enumerate(records, 2):
        ws.cell(row=row_idx, column=1, value=r.athlete.full_name)
        ws.cell(row=row_idx, column=2, value=str(r.attendance_date))
        ws.cell(row=row_idx, column=3, value=r.status)
        ws.cell(row=row_idx, column=4, value=r.session_type)
        ws.cell(row=row_idx, column=5, value=r.notes)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer